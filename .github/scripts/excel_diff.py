# .github/scripts/excel_diff.py
"""
Compare Excel files between base and head branch of a PR.

Two layers:
- ``compute_diff(old_data, new_data)`` returns a structured dict describing
  per-sheet changes (sheet add/remove, row add/delete/change, column
  add/remove). This is the input the version-bump computation consumes.
- ``render_markdown(diff)`` consumes that dict and produces the PR-comment
  Markdown.

Usage:
    python excel_diff.py --files "path/to/file.xlsx" --base-ref main --output diff.md
"""

import argparse
import os
import subprocess
import sys
import tempfile

from openpyxl import load_workbook


def extract_workbook(path: str) -> dict[str, list[list[str]]]:
    """Extract all sheets into a dict of {sheet_name: [[cell_values]]}."""
    wb = load_workbook(path, data_only=True)
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows():
            rows.append([str(c.value) if c.value is not None else "" for c in row])
        data[sheet_name] = rows
    return data


def get_workbook_at_ref(filepath: str, ref: str) -> dict[str, list[list[str]]]:
    """Extract workbook content at an arbitrary git ref (branch, tag, sha)."""
    try:
        result = subprocess.run(
            ["git", "show", f"{ref}:{filepath}"],
            capture_output=True,
        )
        if result.returncode != 0:
            return {}

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(result.stdout)
        tmp.close()
        data = extract_workbook(tmp.name)
        os.unlink(tmp.name)
        return data
    except Exception:
        return {}


def get_base_version(filepath: str, base_ref: str) -> dict[str, list[list[str]]]:
    """Extract the base branch version of a file via git show origin/<ref>."""
    return get_workbook_at_ref(filepath, f"origin/{base_ref}")


def find_id_column(headers: list[str]) -> int | None:
    """Find the index of the 'id' column (case-insensitive)."""
    for i, h in enumerate(headers):
        if h.strip().lower() == "id":
            return i
    return None


def find_duplicates(rows: list[list[str]], id_col: int) -> list[str]:
    """Return list of duplicate ID values."""
    seen = {}
    duplicates = []
    for row in rows:
        if id_col < len(row):
            val = row[id_col]
            if val in seen:
                if val not in duplicates:
                    duplicates.append(val)
            else:
                seen[val] = True
    return duplicates


def build_row_index(rows: list[list[str]], id_col: int) -> dict[str, list[str]]:
    """Build a dict mapping ID value -> row data. Last occurrence wins."""
    index = {}
    for row in rows:
        if id_col < len(row):
            index[row[id_col]] = row
    return index


def _changed_column_names(
    old_headers: list[str],
    new_headers: list[str],
    old_row: list[str],
    new_row: list[str],
) -> list[str]:
    """Return the names of columns whose value differs between old_row and new_row.

    Compares only on the intersection of headers — column add/remove is
    reported separately at the sheet level.
    """
    common = [h for h in new_headers if h in old_headers]
    changed = []
    for h in common:
        oi = old_headers.index(h)
        ni = new_headers.index(h)
        ov = old_row[oi] if oi < len(old_row) else ""
        nv = new_row[ni] if ni < len(new_row) else ""
        if ov != nv:
            changed.append(h)
    return changed


def _compute_sheet_diff_id(
    old_rows: list[list[str]],
    new_rows: list[list[str]],
) -> dict:
    """Compute structured diff for an ID-keyed sheet.

    Caller must have determined that the (header) sheet has an ``id`` column.
    """
    old_headers = old_rows[0] if old_rows else []
    new_headers = new_rows[0] if new_rows else []
    headers_for_render = old_headers or new_headers

    id_col_old = find_id_column(old_headers) if old_headers else None
    id_col_new = find_id_column(new_headers) if new_headers else None

    old_data = old_rows[1:]
    new_data = new_rows[1:]

    old_dupes = (
        find_duplicates(old_data, id_col_old) if id_col_old is not None else []
    )
    new_dupes = (
        find_duplicates(new_data, id_col_new) if id_col_new is not None else []
    )

    if old_dupes or new_dupes:
        return {
            "event": "error",
            "error_type": "duplicate_ids",
            "old_duplicates": old_dupes,
            "new_duplicates": new_dupes,
        }

    columns_added = [h for h in new_headers if h not in old_headers]
    columns_removed = [h for h in old_headers if h not in new_headers]

    old_index = (
        build_row_index(old_data, id_col_old) if id_col_old is not None else {}
    )
    new_index = (
        build_row_index(new_data, id_col_new) if id_col_new is not None else {}
    )

    # Order: ids in the new sheet first, then ids only in old (deletions)
    seen = set()
    ordered_ids = []
    for k in new_index:
        ordered_ids.append(k)
        seen.add(k)
    for k in old_index:
        if k not in seen:
            ordered_ids.append(k)

    row_changes: list[dict] = []
    for id_val in ordered_ids:
        in_old = id_val in old_index
        in_new = id_val in new_index
        if in_old and not in_new:
            row_changes.append(
                {"type": "deleted", "key": id_val, "row": old_index[id_val]}
            )
        elif not in_old and in_new:
            row_changes.append(
                {"type": "added", "key": id_val, "row": new_index[id_val]}
            )
        else:
            old_row = old_index[id_val]
            new_row = new_index[id_val]
            if old_row != new_row:
                changed_columns = _changed_column_names(
                    old_headers, new_headers, old_row, new_row
                )
                id_changed = (
                    id_col_old is not None
                    and id_col_new is not None
                    and id_col_old < len(old_row)
                    and id_col_new < len(new_row)
                    and old_row[id_col_old] != new_row[id_col_new]
                )
                row_changes.append(
                    {
                        "type": "changed",
                        "key": id_val,
                        "old_row": old_row,
                        "new_row": new_row,
                        "changed_columns": changed_columns,
                        "id_changed": id_changed,
                    }
                )

    if not (row_changes or columns_added or columns_removed):
        return {"event": "unchanged"}

    return {
        "event": "modified",
        "mode": "id",
        "headers": headers_for_render,
        "columns_added": columns_added,
        "columns_removed": columns_removed,
        "row_changes": row_changes,
    }


def _compute_sheet_diff_positional(
    old_rows: list[list[str]],
    new_rows: list[list[str]],
) -> dict:
    """Fallback: positional diff when no ID column exists."""
    old_headers = old_rows[0] if old_rows else []
    new_headers = new_rows[0] if new_rows else []
    headers_for_render = old_headers or new_headers

    columns_added = [h for h in new_headers if h not in old_headers]
    columns_removed = [h for h in old_headers if h not in new_headers]

    row_changes: list[dict] = []
    max_rows = max(len(old_rows), len(new_rows))
    for r in range(1, max_rows):
        old_row = old_rows[r] if r < len(old_rows) else None
        new_row = new_rows[r] if r < len(new_rows) else None
        if old_row is None and new_row is not None:
            row_changes.append({"type": "added", "key": r, "row": new_row})
        elif old_row is not None and new_row is None:
            row_changes.append({"type": "deleted", "key": r, "row": old_row})
        elif old_row != new_row:
            changed_columns = _changed_column_names(
                old_headers, new_headers, old_row, new_row
            )
            row_changes.append(
                {
                    "type": "changed",
                    "key": r,
                    "old_row": old_row,
                    "new_row": new_row,
                    "changed_columns": changed_columns,
                    "id_changed": False,
                }
            )

    if not (row_changes or columns_added or columns_removed):
        return {"event": "unchanged"}

    return {
        "event": "modified",
        "mode": "positional",
        "headers": headers_for_render,
        "columns_added": columns_added,
        "columns_removed": columns_removed,
        "row_changes": row_changes,
    }


def compute_diff(
    old_data: dict[str, list[list[str]]],
    new_data: dict[str, list[list[str]]],
) -> dict:
    """Compute a structured per-sheet diff.

    Returns ``{"sheets": {<name>: <sheet_diff>}, "has_errors": bool}``.

    Sheet diff events:
      - ``added``    — sheet exists in new but not old
      - ``removed``  — sheet exists in old but not new
      - ``unchanged`` — present on both sides, no differences
      - ``error``    — validation error (e.g. duplicate IDs); ``error_type`` set
      - ``modified`` — present on both sides with differences. Carries
        ``mode`` (``id`` or ``positional``), ``headers``, ``columns_added``,
        ``columns_removed``, and ``row_changes`` (ordered list of
        ``{type, key, ...}`` entries).
    """
    sheets: dict[str, dict] = {}
    has_errors = False

    all_sheets = sorted(set(list(old_data.keys()) + list(new_data.keys())))
    for sheet in all_sheets:
        if sheet not in old_data:
            sheets[sheet] = {"event": "added"}
            continue
        if sheet not in new_data:
            sheets[sheet] = {"event": "removed"}
            continue

        old_rows = old_data[sheet]
        new_rows = new_data[sheet]

        headers = old_rows[0] if old_rows else new_rows[0] if new_rows else []
        if not headers:
            sheets[sheet] = {"event": "unchanged"}
            continue

        id_col = find_id_column(headers)
        if id_col is not None:
            sd = _compute_sheet_diff_id(old_rows, new_rows)
        else:
            sd = _compute_sheet_diff_positional(old_rows, new_rows)

        sheets[sheet] = sd
        if sd.get("event") == "error":
            has_errors = True

    return {"sheets": sheets, "has_errors": has_errors}


def highlight_changed_cells(
    headers: list[str], old_row: list[str], new_row: list[str]
) -> tuple[str, str]:
    """Format old and new rows, bolding cells that changed."""
    max_cols = len(headers)
    old_padded = old_row + [""] * (max_cols - len(old_row))
    new_padded = new_row + [""] * (max_cols - len(new_row))

    old_cells = []
    new_cells = []
    for i in range(max_cols):
        old_val = old_padded[i].replace("|", "\\|")
        new_val = new_padded[i].replace("|", "\\|")
        if old_val != new_val:
            old_cells.append(f"**{old_val}**")
            new_cells.append(f"**{new_val}**")
        else:
            old_cells.append(old_val)
            new_cells.append(new_val)

    return (
        "| " + " | ".join(old_cells) + " |",
        "| " + " | ".join(new_cells) + " |",
    )


def format_row(headers: list[str], row: list[str]) -> str:
    """Format a single data row as a Markdown table row."""
    padded = row + [""] * (len(headers) - len(row))
    return (
        "| " + " | ".join(v.replace("|", "\\|") for v in padded[: len(headers)]) + " |"
    )


def _render_modified_sheet(sheet: str, sd: dict) -> list[str]:
    """Render a single 'modified' sheet diff as Markdown lines."""
    lines: list[str] = []
    headers = sd["headers"]
    row_changes = sd["row_changes"]
    columns_added = sd.get("columns_added", [])
    columns_removed = sd.get("columns_removed", [])
    suffix = "" if sd["mode"] == "id" else " (no ID column, positional diff)"

    if row_changes:
        lines.append(
            f"### 📝 Sheet: `{sheet}` — {len(row_changes)} row(s) affected{suffix}"
        )
    else:
        lines.append(f"### 📝 Sheet: `{sheet}` — column changes only{suffix}")
    lines.append("")

    if columns_added:
        lines.append(f"**🟢 Columns added:** `{'`, `'.join(columns_added)}`")
        lines.append("")
    if columns_removed:
        lines.append(f"**🔴 Columns removed:** `{'`, `'.join(columns_removed)}`")
        lines.append("")

    if not row_changes:
        return lines

    header_line = "| " + " | ".join(h.replace("|", "\\|") for h in headers) + " |"
    sep_line = "| " + " | ".join("---" for _ in headers) + " |"

    for change in row_changes:
        ctype = change["type"]
        key = change["key"]
        if ctype == "added":
            label = (
                f"**Row `{key}` — Added**"
                if sd["mode"] == "id"
                else f"**Row {key} — Added**"
            )
            lines.append(label)
            lines.append("")
            lines.append(header_line)
            lines.append(sep_line)
            lines.append(format_row(headers, change["row"]))
            lines.append("")
        elif ctype == "deleted":
            label = (
                f"**⚠️ Row `{key}` — Deleted**"
                if sd["mode"] == "id"
                else f"**⚠️ Row {key} — Deleted**"
            )
            lines.append(label)
            lines.append("")
            lines.append(header_line)
            lines.append(sep_line)
            lines.append(format_row(headers, change["row"]))
            lines.append("")
        elif ctype == "changed":
            if change.get("id_changed"):
                label = f"**⚠️ Row `{key}` — Key change**"
            else:
                if sd["mode"] == "id":
                    label = f"**Row `{key}` — Changed** (changed cells in bold)"
                else:
                    label = f"**Row {key} — Changed** (changed cells in bold)"
            old_line, new_line = highlight_changed_cells(
                headers, change["old_row"], change["new_row"]
            )
            lines.append(label)
            lines.append("")
            lines.append(header_line)
            lines.append(sep_line)
            lines.append(old_line)
            lines.append(new_line)
            lines.append("")

    return lines


def render_markdown(diff: dict) -> str:
    """Render a structured diff as the PR-comment Markdown report."""
    lines: list[str] = []
    for sheet, sd in diff["sheets"].items():
        event = sd["event"]
        if event == "added":
            lines.append(f"### 🟢 Sheet added: `{sheet}`")
            lines.append("")
        elif event == "removed":
            lines.append(f"### 🔴 Sheet removed: `{sheet}`")
            lines.append("")
        elif event == "unchanged":
            continue
        elif event == "error":
            if sd.get("error_type") == "duplicate_ids":
                lines.append(f"### ❌ Sheet: `{sheet}` — Duplicate IDs detected")
                lines.append("")
                if sd["old_duplicates"]:
                    lines.append(
                        f"Base branch duplicates: `{'`, `'.join(sd['old_duplicates'])}`"
                    )
                if sd["new_duplicates"]:
                    lines.append(
                        f"PR branch duplicates: `{'`, `'.join(sd['new_duplicates'])}`"
                    )
                lines.append("")
                lines.append("Resolve duplicate IDs before meaningful diff is possible.")
                lines.append("")
        elif event == "modified":
            lines.extend(_render_modified_sheet(sheet, sd))
    return "\n".join(lines)


def diff_workbooks(
    old_data: dict[str, list[list[str]]],
    new_data: dict[str, list[list[str]]],
) -> str:
    """Backward-compatible wrapper — compute diff and render to Markdown."""
    return render_markdown(compute_diff(old_data, new_data))


def main():
    parser = argparse.ArgumentParser(
        description="Diff Excel files between git branches"
    )
    parser.add_argument(
        "--files", required=True, help="Newline-separated list of xlsx paths"
    )
    parser.add_argument(
        "--base-ref", required=True, help="Base branch name (e.g. main)"
    )
    parser.add_argument("--output", required=True, help="Output markdown file path")
    args = parser.parse_args()

    files = [f.strip() for f in args.files.strip().splitlines() if f.strip()]
    all_output = []

    for filepath in files:
        all_output.append(f"## 📊 `{filepath}`\n")
        try:
            old_data = get_base_version(filepath, args.base_ref)
            new_data = extract_workbook(filepath)
            all_output.append(diff_workbooks(old_data, new_data))
        except Exception as e:
            all_output.append(f"⚠️ Error processing file: {e}\n")

    output = "\n".join(all_output)
    has_errors = "❌" in output

    with open(args.output, "w") as f:
        f.write(output)

    print(f"Generated diff for {len(files)} file(s) -> {args.output}")

    if has_errors:
        print("::error::Validation errors found in Excel file(s)")
        sys.exit(1)


if __name__ == "__main__":
    main()
