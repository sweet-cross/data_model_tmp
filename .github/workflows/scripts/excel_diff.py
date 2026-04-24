# .github/scripts/excel_diff.py
"""
Compare Excel files between base and head branch of a PR.
Outputs a Markdown diff report.

Usage:
    python excel_diff.py --files "path/to/file.xlsx" --base-ref main --output diff.md
"""

import argparse
import os
import subprocess
import tempfile

from openpyxl import load_workbook


def extract_workbook(path: str) -> dict[str, list[list[str]]]:
    """Extract all sheets into a dict of {sheet_name: [[cell_values]]}."""
    wb = load_workbook(path, data_only=True)
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows():
            rows.append([str(c.value) if c.value is not None else "" for c in row])
        data[sheet_name] = rows
    return data


def get_base_version(filepath: str, base_ref: str) -> dict[str, list[list[str]]]:
    """Extract the base branch version of a file via git show."""
    try:
        result = subprocess.run(
            ["git", "show", f"origin/{base_ref}:{filepath}"],
            capture_output=True,
        )
        if result.returncode != 0:
            return {}  # file is new in this PR

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(result.stdout)
        tmp.close()
        data = extract_workbook(tmp.name)
        os.unlink(tmp.name)
        return data
    except Exception:
        return {}


def col_index_to_letter(index: int) -> str:
    """Convert a 0-based column index to Excel-style letter (0=A, 25=Z, 26=AA)."""
    col_letter = ""
    col_num = index
    while col_num >= 0:
        col_letter = chr(col_num % 26 + 65) + col_letter
        col_num = col_num // 26 - 1
    return col_letter


def diff_workbooks(
    old_data: dict[str, list[list[str]]],
    new_data: dict[str, list[list[str]]],
    max_changes_per_sheet: int = 200,
) -> str:
    """Compare two workbook extracts, return a Markdown diff."""
    lines = []
    all_sheets = sorted(set(list(old_data.keys()) + list(new_data.keys())))

    for sheet in all_sheets:
        if sheet not in old_data:
            lines.append(f"### 🟢 Sheet added: `{sheet}`")
            lines.append("")
            continue
        if sheet not in new_data:
            lines.append(f"### 🔴 Sheet removed: `{sheet}`")
            lines.append("")
            continue

        old_rows = old_data[sheet]
        new_rows = new_data[sheet]
        changes = []

        max_rows = max(len(old_rows), len(new_rows))
        for r in range(max_rows):
            old_row = old_rows[r] if r < len(old_rows) else []
            new_row = new_rows[r] if r < len(new_rows) else []
            max_cols = max(len(old_row), len(new_row))

            for c in range(max_cols):
                old_val = old_row[c] if c < len(old_row) else ""
                new_val = new_row[c] if c < len(new_row) else ""
                if old_val != new_val:
                    cell_ref = f"{col_index_to_letter(c)}{r + 1}"
                    changes.append((cell_ref, old_val, new_val))

        if changes:
            lines.append(f"### 📝 Sheet: `{sheet}` — {len(changes)} change(s)")
            lines.append("")
            lines.append("| Cell | Old Value | New Value |")
            lines.append("|------|-----------|-----------|")
            for cell, old, new in changes[:max_changes_per_sheet]:
                old_esc = old.replace("|", "\\|")
                new_esc = new.replace("|", "\\|")
                lines.append(f"| `{cell}` | {old_esc} | {new_esc} |")
            if len(changes) > max_changes_per_sheet:
                remaining = len(changes) - max_changes_per_sheet
                lines.append(f"| ... | _{remaining} more changes_ | |")
            lines.append("")
        else:
            lines.append(f"### ✅ Sheet: `{sheet}` — no changes")
            lines.append("")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Diff Excel files between git branches"
    )
    parser.add_argument(
        "--files", required=True, help="Newline-separated list of xlsx paths"
    )
    parser.add_argument(
        "--base-ref", required=True, help="Base branch name (e.g. main)"
    )
    parser.add_argument("--output", required=True, help="Output markdown file path")
    args = parser.parse_args()

    files = [f.strip() for f in args.files.strip().splitlines() if f.strip()]
    all_output = []

    for filepath in files:
        all_output.append(f"## 📊 `{filepath}`\n")
        try:
            old_data = get_base_version(filepath, args.base_ref)
            new_data = extract_workbook(filepath)
            all_output.append(diff_workbooks(old_data, new_data))
        except Exception as e:
            all_output.append(f"⚠️ Error processing file: {e}\n")

    output = "\n".join(all_output)

    with open(args.output, "w") as f:
        f.write(output)

    print(f"Generated diff for {len(files)} file(s) -> {args.output}")


if __name__ == "__main__":
    main()
